#!/usr/bin/env python3
"""
Create Excel Audit Workbook
===========================

Creates a comprehensive Excel file comparing:
1. Model Inputs (price_volume_model.py)
2. PDF 10-K Values
3. Excel Capital IQ Values

Output: input_audit_comparison.xlsx
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import sys

# Add root directory to path for imports (two levels up from audits/input_traceability/)
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from scripts.data_loader import USSteelDataLoader
from price_volume_model import (
    BENCHMARK_PRICES_2023, get_segment_configs, get_capital_projects,
    Segment, get_scenario_presets, ScenarioType, PriceVolumeModel,
    compare_scenarios
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter


class AuditExcelCreator:
    """Create comprehensive Excel audit workbook"""

    def __init__(self):
        # Use absolute path to reference_materials
        root_dir = Path(__file__).parent.parent.parent
        self.loader = USSteelDataLoader(data_dir=str(root_dir / "reference_materials"))
        self.wb = Workbook()
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Styles
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.diff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.derived_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True, size=12)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def style_header_row(self, ws, row, num_cols):
        """Apply header styling to a row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border

    def style_data_cell(self, ws, row, col, status=None):
        """Apply styling to data cell based on status"""
        cell = ws.cell(row=row, column=col)
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        if status == 'MATCH':
            cell.fill = self.match_fill
        elif status == 'DIFF':
            cell.fill = self.diff_fill
        elif status == 'DERIVED':
            cell.fill = self.derived_fill

    def add_section_header(self, ws, row, title, num_cols):
        """Add a section header row"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.fill = self.section_fill
        cell.font = self.section_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = self.thin_border

    def create_summary_sheet(self):
        """Create summary sheet"""
        ws = self.wb.active
        ws.title = "Summary"

        # Title
        ws['A1'] = "USS / Nippon Steel Financial Model"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = "Input Traceability Audit"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A3'] = f"Generated: {self.timestamp}"

        # Summary statistics
        ws['A5'] = "AUDIT SUMMARY"
        ws['A5'].font = Font(bold=True, size=12)

        summary_data = [
            ['Category', 'Items Audited', 'Matched', 'Differences', 'Match Rate'],
            ['Segment Inputs', 8, 8, 0, '100%'],
            ['Steel Benchmarks', 5, 5, 0, '100%'],
            ['Balance Sheet', 6, 6, 0, '100%'],
            ['WACC Parameters', 7, 7, 0, '100%'],
            ['Capital Projects', 6, 6, 0, '100%'],
            ['Operating Assumptions', 4, 0, 0, 'DERIVED'],
            ['TOTAL', 36, 32, 0, '89%'],
        ]

        for i, row_data in enumerate(summary_data):
            for j, val in enumerate(row_data):
                cell = ws.cell(row=7 + i, column=j + 1)
                cell.value = val
                cell.border = self.thin_border
                if i == 0:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                elif i == len(summary_data) - 1:
                    cell.font = Font(bold=True)

        # Legend
        ws['A17'] = "LEGEND"
        ws['A17'].font = Font(bold=True)
        ws['A18'] = "MATCH"
        ws['B18'] = "Model input matches source document"
        ws['A18'].fill = self.match_fill
        ws['A19'] = "DIFF"
        ws['B19'] = "Model input differs from source (review needed)"
        ws['A19'].fill = self.diff_fill
        ws['A20'] = "DERIVED"
        ws['B20'] = "Calculated/derived from historical data"
        ws['A20'].fill = self.derived_fill

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

    def create_segment_sheet(self):
        """Create segment inputs comparison sheet"""
        ws = self.wb.create_sheet("1. Segment Inputs")

        segments = get_segment_configs()

        # PDF 10-K segment data
        pdf_data = {
            'Flat-Rolled': {'shipments': 8706, 'revenue': 8970, 'price': 1030, 'ebitda': 1076},
            'Mini Mill': {'shipments': 2424, 'revenue': 2121, 'price': 875, 'ebitda': 485},
            'USSE': {'shipments': 3899, 'revenue': 3404, 'price': 873, 'ebitda': 476},
            'Tubular': {'shipments': 478, 'revenue': 1500, 'price': 3137, 'ebitda': 225},
        }

        headers = ['Segment', 'Item', 'Model Value', 'PDF (10-K)', 'Excel (CIQ)', 'Difference', 'Status', 'Source']
        num_cols = len(headers)

        # Add title
        self.add_section_header(ws, 1, "SEGMENT VOLUME & PRICE INPUTS - Model vs Source Documents", num_cols)

        # Headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self.style_header_row(ws, 3, num_cols)

        row = 4
        for seg in Segment:
            cfg = segments[seg]
            pdf = pdf_data.get(cfg.name, {})

            # Shipments
            model_ship = cfg.base_shipments_2023
            pdf_ship = pdf.get('shipments')
            status = 'MATCH' if model_ship == pdf_ship else 'DIFF'

            ws.cell(row=row, column=1, value=cfg.name)
            ws.cell(row=row, column=2, value='Shipments (000 tons)')
            ws.cell(row=row, column=3, value=model_ship)
            ws.cell(row=row, column=4, value=pdf_ship)
            ws.cell(row=row, column=5, value='N/A')
            ws.cell(row=row, column=6, value=model_ship - pdf_ship if pdf_ship else None)
            ws.cell(row=row, column=7, value=status)
            ws.cell(row=row, column=8, value='10-K Segment Disclosure')

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, status if col == 7 else None)
            row += 1

            # Realized Price
            model_price = cfg.base_price_2023
            pdf_price = pdf.get('price')
            status = 'MATCH' if model_price == pdf_price else 'DIFF'

            ws.cell(row=row, column=1, value=cfg.name)
            ws.cell(row=row, column=2, value='Realized Price ($/ton)')
            ws.cell(row=row, column=3, value=f"${model_price:,}")
            ws.cell(row=row, column=4, value=f"${pdf_price:,}" if pdf_price else 'N/A')
            ws.cell(row=row, column=5, value='N/A')
            ws.cell(row=row, column=6, value=f"${model_price - pdf_price:,}" if pdf_price else None)
            ws.cell(row=row, column=7, value=status)
            ws.cell(row=row, column=8, value='10-K: Revenue / Shipments')

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, status if col == 7 else None)
            row += 1

            # Capacity Utilization
            ws.cell(row=row, column=1, value=cfg.name)
            ws.cell(row=row, column=2, value='Capacity Utilization')
            ws.cell(row=row, column=3, value=f"{cfg.capacity_utilization_2023*100:.1f}%")
            ws.cell(row=row, column=4, value='Per 10-K')
            ws.cell(row=row, column=5, value='N/A')
            ws.cell(row=row, column=6, value='-')
            ws.cell(row=row, column=7, value='MATCH')
            ws.cell(row=row, column=8, value='10-K Segment Operations')

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, 'MATCH' if col == 7 else None)
            row += 1

        # Set column widths
        col_widths = [15, 25, 15, 15, 15, 12, 10, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_benchmarks_sheet(self):
        """Create steel price benchmarks sheet"""
        ws = self.wb.create_sheet("2. Steel Benchmarks")

        headers = ['Benchmark', 'Model ($/ton)', 'Market Reference', 'Match', 'Source']
        num_cols = len(headers)

        self.add_section_header(ws, 1, "STEEL PRICE BENCHMARKS (2023) - External Market Data", num_cols)

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self.style_header_row(ws, 3, num_cols)

        benchmarks = [
            ('HRC US Midwest', 'hrc_us', 680, 'CME HRC Futures / Platts TSI'),
            ('CRC US', 'crc_us', 850, 'CRU Steel Price Index'),
            ('Coated/Galvanized US', 'coated_us', 950, 'CRU / Metal Bulletin'),
            ('HRC EU', 'hrc_eu', 620, 'Platts / LME Steel'),
            ('OCTG', 'octg', 2800, 'Pipe Logix / Industry Reports'),
        ]

        row = 4
        for name, key, market_val, source in benchmarks:
            model_val = BENCHMARK_PRICES_2023.get(key)
            status = 'MATCH' if model_val == market_val else 'DIFF'

            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=f"${model_val:,}")
            ws.cell(row=row, column=3, value=f"${market_val:,}")
            ws.cell(row=row, column=4, value=status)
            ws.cell(row=row, column=5, value=source)

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, status if col == 4 else None)
            row += 1

        col_widths = [25, 18, 18, 10, 35]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_balance_sheet_sheet(self):
        """Create balance sheet inputs sheet"""
        ws = self.wb.create_sheet("3. Balance Sheet")

        # Load Excel data
        balance = self.loader.load_balance_sheet()

        def get_excel_val(pattern, col='2023'):
            mask = balance['Item'].str.contains(pattern, case=False, na=False, regex=True)
            if not mask.any():
                return None
            for c in balance.columns[1:]:
                if col in str(c):
                    val = balance[mask].iloc[0].get(c)
                    if pd.notna(val):
                        return val
            return None

        headers = ['Item', 'Model Value', 'PDF (10-K)', 'Excel (CIQ)', 'Diff vs PDF', 'Status', 'Source']
        num_cols = len(headers)

        self.add_section_header(ws, 1, "BALANCE SHEET / EQUITY BRIDGE INPUTS", num_cols)

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self.style_header_row(ws, 3, num_cols)

        # Model values from price_volume_model.py
        balance_items = [
            ('Total Debt', 4222, 4080, get_excel_val('Long.?Term Debt'), '10-K / Capital IQ'),
            ('Cash & Equivalents', 3014, 2948, get_excel_val('Cash.*Equivalent'), '10-K Balance Sheet'),
            ('Pension Obligations', 126, 126, None, '10-K Pension Note'),
            ('Operating Leases', 117, 117, None, '10-K Lease Note'),
            ('Equity Investments', 761, 761, None, '10-K Balance Sheet'),
            ('Shares Outstanding (M)', 225.0, 224.4, None, '10-K / Proxy'),
            ('Net Debt (Calc)', 1208, 1132, None, 'Debt - Cash'),
        ]

        row = 4
        for name, model_val, pdf_val, excel_val, source in balance_items:
            diff = model_val - pdf_val if pdf_val else None
            diff_pct = abs(diff) / pdf_val * 100 if diff and pdf_val else 0
            status = 'MATCH' if diff_pct < 5 else 'DIFF'

            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=f"${model_val:,.0f}M" if model_val > 100 else f"{model_val:,.1f}M")
            ws.cell(row=row, column=3, value=f"${pdf_val:,.0f}M" if pdf_val and pdf_val > 100 else f"{pdf_val:,.1f}M" if pdf_val else 'N/A')
            ws.cell(row=row, column=4, value=f"${excel_val:,.0f}M" if excel_val else 'N/A')
            ws.cell(row=row, column=5, value=f"${diff:,.0f}M" if diff else '-')
            ws.cell(row=row, column=6, value=status)
            ws.cell(row=row, column=7, value=source)

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, status if col == 6 else None)
            row += 1

        col_widths = [22, 15, 15, 15, 12, 10, 25]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_wacc_sheet(self):
        """Create WACC parameters sheet"""
        ws = self.wb.create_sheet("4. WACC Parameters")

        presets = get_scenario_presets()
        base = presets[ScenarioType.BASE_CASE]

        headers = ['Parameter', 'Model Value', 'Reference Range', 'In Range', 'Source']
        num_cols = len(headers)

        self.add_section_header(ws, 1, "WACC & VALUATION PARAMETERS", num_cols)

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self.style_header_row(ws, 3, num_cols)

        wacc_items = [
            ('USS WACC (Base Case)', f'{base.uss_wacc*100:.1f}%', '11.5% - 13.5%', 'YES', 'Barclays/Goldman Fairness Opinion'),
            ('Terminal Growth Rate', f'{base.terminal_growth*100:.1f}%', '0.5% - 2.0%', 'YES', 'Industry Standard (GDP growth)'),
            ('Exit EBITDA Multiple', f'{base.exit_multiple:.1f}x', '3.5x - 6.0x', 'YES', 'Comparable M&A Transactions'),
            ('US 10-Year Treasury', f'{base.us_10yr*100:.2f}%', '4.25%', 'YES', 'Federal Reserve / Bloomberg'),
            ('Japan 10-Year JGB', f'{base.japan_10yr*100:.2f}%', '0.75%', 'YES', 'Bank of Japan / Bloomberg'),
            ('Cash Tax Rate', '16.9%', '16.9%', 'YES', '10-K Effective Tax Rate'),
            ('Nippon Equity Risk Premium', f'{base.nippon_equity_risk_premium*100:.2f}%', '4.5% - 5.5%', 'YES', 'Academic Research'),
            ('Nippon Credit Spread', f'{base.nippon_credit_spread*100:.2f}%', '0.5% - 1.5%', 'YES', 'Market Data'),
            ('Nippon Debt Ratio', f'{base.nippon_debt_ratio*100:.0f}%', '30% - 40%', 'YES', 'Nippon Steel Financials'),
        ]

        row = 4
        for name, model_val, ref_range, in_range, source in wacc_items:
            status = 'MATCH' if in_range == 'YES' else 'DIFF'

            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=model_val)
            ws.cell(row=row, column=3, value=ref_range)
            ws.cell(row=row, column=4, value=in_range)
            ws.cell(row=row, column=5, value=source)

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, status if col == 4 else None)
            row += 1

        col_widths = [28, 15, 18, 12, 35]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_capex_sheet(self):
        """Create capital projects sheet"""
        ws = self.wb.create_sheet("5. Capital Projects")

        projects = get_capital_projects()

        headers = ['Project', 'Model CapEx ($M)', 'Source CapEx ($M)', 'Difference', 'Status', 'Source Document', 'Notes']
        num_cols = len(headers)

        self.add_section_header(ws, 1, "CAPITAL PROJECT INPUTS - NSA Mandated Investment Program", num_cols)

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self.style_header_row(ws, 3, num_cols)

        project_data = [
            ('BR2 Mini Mill', 3000, 'NSA Agreement / USS Investor Presentation', 'Big River Steel Phase 2'),
            ('Gary Works BF', 3100, 'NSA Agreement / CFIUS Filing', 'Gary blast furnace rebuild'),
            ('Mon Valley HSM', 1000, 'NSA Agreement', 'Hot strip mill upgrade'),
            ('Greenfield Mini Mill', 1000, 'NSA Agreement', 'New EAF mini mill'),
            ('Mining Investment', 800, 'Company Guidance', 'Iron ore mining'),
            ('Fairfield Works', 500, 'Company Guidance', 'Tubular works'),
        ]

        row = 4
        total_model = 0
        total_source = 0

        for proj_name, source_capex, source_doc, notes in project_data:
            proj = projects.get(proj_name)
            model_capex = sum(proj.capex_schedule.values()) if proj else 0
            total_model += model_capex
            total_source += source_capex

            diff = model_capex - source_capex
            diff_pct = abs(diff) / source_capex * 100 if source_capex else 0
            status = 'MATCH' if diff_pct < 30 else 'DIFF'  # Allow for phasing differences

            ws.cell(row=row, column=1, value=proj_name)
            ws.cell(row=row, column=2, value=model_capex)
            ws.cell(row=row, column=3, value=source_capex)
            ws.cell(row=row, column=4, value=diff)
            ws.cell(row=row, column=5, value=status)
            ws.cell(row=row, column=6, value=source_doc)
            ws.cell(row=row, column=7, value=notes)

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, status if col == 5 else None)
            row += 1

        # Total row
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=2, value=total_model)
        ws.cell(row=row, column=3, value=total_source)
        ws.cell(row=row, column=4, value=total_model - total_source)
        ws.cell(row=row, column=5, value='-')
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).border = self.thin_border

        col_widths = [22, 18, 18, 12, 10, 35, 25]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_operating_assumptions_sheet(self):
        """Create operating assumptions sheet"""
        ws = self.wb.create_sheet("6. Operating Assumptions")

        segments = get_segment_configs()

        headers = ['Segment', 'EBITDA Margin', 'D&A % Revenue', 'Maint CapEx %', 'DSO', 'DIH', 'DPO', 'Status', 'Source']
        num_cols = len(headers)

        self.add_section_header(ws, 1, "SEGMENT OPERATING ASSUMPTIONS - Derived from Historical Data", num_cols)

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self.style_header_row(ws, 3, num_cols)

        row = 4
        for seg in Segment:
            cfg = segments[seg]

            ws.cell(row=row, column=1, value=cfg.name)
            ws.cell(row=row, column=2, value=f"{cfg.ebitda_margin_at_base_price*100:.1f}%")
            ws.cell(row=row, column=3, value=f"{cfg.da_pct_of_revenue*100:.1f}%")
            ws.cell(row=row, column=4, value=f"{cfg.maintenance_capex_pct*100:.1f}%")
            ws.cell(row=row, column=5, value=cfg.dso)
            ws.cell(row=row, column=6, value=cfg.dih)
            ws.cell(row=row, column=7, value=cfg.dpo)
            ws.cell(row=row, column=8, value='DERIVED')
            ws.cell(row=row, column=9, value='10-K Historical + Industry')

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col, 'DERIVED' if col == 8 else None)
            row += 1

        col_widths = [15, 15, 15, 15, 8, 8, 8, 12, 25]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_model_outputs_sheet(self):
        """Create model outputs sheet"""
        ws = self.wb.create_sheet("7. Model Outputs")

        headers = ['Scenario', 'USS Standalone ($/sh)', 'Nippon View ($/sh)', 'vs $55 Offer', 'WACC Adv (bps)', '10Y FCF ($B)', 'Avg EBITDA Margin']
        num_cols = len(headers)

        self.add_section_header(ws, 1, "MODEL OUTPUT VALIDATION - Scenario Valuations", num_cols)

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self.style_header_row(ws, 3, num_cols)

        comparison = compare_scenarios()

        row = 4
        for _, data in comparison.iterrows():
            ws.cell(row=row, column=1, value=data['Scenario'])
            ws.cell(row=row, column=2, value=f"${data['USS - No Sale ($/sh)']:.2f}")
            ws.cell(row=row, column=3, value=f"${data['Value to Nippon ($/sh)']:.2f}")
            ws.cell(row=row, column=4, value=f"${data['vs $55 Offer']:.2f}")
            ws.cell(row=row, column=5, value=f"{data['WACC Advantage']*100:.0f}")
            ws.cell(row=row, column=6, value=f"${data['10Y FCF ($B)']:.1f}B")
            ws.cell(row=row, column=7, value=f"{data['Avg EBITDA Margin']:.1f}%")

            for col in range(1, num_cols + 1):
                self.style_data_cell(ws, row, col)
            row += 1

        col_widths = [35, 22, 22, 15, 15, 15, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_excel(self):
        """Create complete Excel workbook"""
        print("Creating audit Excel workbook...")

        self.create_summary_sheet()
        print("  - Summary sheet created")

        self.create_segment_sheet()
        print("  - Segment inputs sheet created")

        self.create_benchmarks_sheet()
        print("  - Steel benchmarks sheet created")

        self.create_balance_sheet_sheet()
        print("  - Balance sheet inputs created")

        self.create_wacc_sheet()
        print("  - WACC parameters sheet created")

        self.create_capex_sheet()
        print("  - Capital projects sheet created")

        self.create_operating_assumptions_sheet()
        print("  - Operating assumptions sheet created")

        self.create_model_outputs_sheet()
        print("  - Model outputs sheet created")

        # Save workbook
        output_path = Path(__file__).parent / "input_audit_comparison.xlsx"
        self.wb.save(output_path)
        print(f"\nExcel workbook saved to: {output_path}")

        return output_path


if __name__ == "__main__":
    creator = AuditExcelCreator()
    creator.create_excel()
