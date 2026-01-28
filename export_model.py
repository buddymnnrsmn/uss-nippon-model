#!/usr/bin/env python3
"""
Model Export Module: Excel Export for USS / Nippon Steel Financial Model
=========================================================================

Exports comprehensive Excel workbooks with:
- Cover sheet with key metrics
- Assumptions and parameters
- DCF valuations (USS and Nippon perspectives)
- FCF reconciliation
- Financial statements (consolidated and by segment)
- WACC analysis
- Capital projects
- Scenario comparisons

Uses styling patterns from create_audit_excel.py.
"""

from io import BytesIO
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from price_volume_model import (
    PriceVolumeModel, ModelScenario, ScenarioType,
    get_scenario_presets, get_segment_configs, get_capital_projects,
    compare_scenarios, calculate_probability_weighted_valuation,
    BENCHMARK_PRICES_2023, Segment, get_synergy_presets, SynergyAssumptions
)


class ExcelStyler:
    """Centralized Excel styling for consistent workbook appearance"""

    def __init__(self):
        # Header styles (dark blue background, white font)
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)

        # Section header (light blue background)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True, size=12)

        # Positive values (green background)
        self.positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.positive_font = Font(color="006100")

        # Negative values (red background)
        self.negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.negative_font = Font(color="9C0006")

        # Neutral/derived (yellow background)
        self.derived_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Title styles
        self.title_font = Font(bold=True, size=16)
        self.subtitle_font = Font(bold=True, size=14)

        # Border
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Number formats
        self.currency_format = '$#,##0'
        self.currency_decimal_format = '$#,##0.00'
        self.percent_format = '0.0%'
        self.percent_decimal_format = '0.00%'
        self.number_format = '#,##0'
        self.decimal_format = '#,##0.00'
        self.multiple_format = '0.0"x"'

    def style_header_row(self, ws, row: int, start_col: int, end_col: int):
        """Apply header styling to a row"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border

    def style_section_header(self, ws, row: int, text: str, start_col: int, end_col: int):
        """Create a section header spanning multiple columns"""
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=start_col)
        cell.value = text
        cell.fill = self.section_fill
        cell.font = self.section_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = self.thin_border

    def style_data_cell(self, ws, row: int, col: int, value_type: str = None):
        """Apply styling to a data cell

        Args:
            value_type: 'positive', 'negative', 'derived', or None
        """
        cell = ws.cell(row=row, column=col)
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        if value_type == 'positive':
            cell.fill = self.positive_fill
            cell.font = self.positive_font
        elif value_type == 'negative':
            cell.fill = self.negative_fill
            cell.font = self.negative_font
        elif value_type == 'derived':
            cell.fill = self.derived_fill

    def auto_column_width(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content"""
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            length = max(min_width, min(length + 2, max_width))
            ws.column_dimensions[column_cells[0].column_letter].width = length


class ModelExporter:
    """Export financial model to Excel workbook"""

    def __init__(self, scenario: ModelScenario, execution_factor: float = 1.0,
                 custom_benchmarks: dict = None):
        self.scenario = scenario
        self.execution_factor = execution_factor
        self.custom_benchmarks = custom_benchmarks or BENCHMARK_PRICES_2023
        self.styler = ExcelStyler()
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Run the model
        self.model = PriceVolumeModel(scenario, execution_factor, custom_benchmarks)
        self.analysis = self.model.run_full_analysis()

    def export_single_scenario(self) -> BytesIO:
        """Export single scenario to Excel workbook

        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()

        # Create all sheets
        self._create_cover_sheet(wb)
        self._create_assumptions_sheet(wb)
        self._create_dcf_sheet(wb, 'uss')
        self._create_dcf_sheet(wb, 'nippon')
        self._create_fcf_reconciliation_sheet(wb)
        self._create_consolidated_sheet(wb)

        # Segment sheets
        for segment in Segment:
            self._create_segment_sheet(wb, segment.value)

        self._create_wacc_sheet(wb)
        self._create_projects_sheet(wb)
        self._create_equity_bridge_sheet(wb)
        self._create_financing_sheet(wb)

        # Add synergy sheet if synergies are enabled
        if self.analysis.get('synergy_schedule') is not None:
            self._create_synergy_sheet(wb)

        # Remove default empty sheet if it exists
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_multi_scenario(self, scenario_types: List[ScenarioType] = None) -> BytesIO:
        """Export multi-scenario comparison to Excel workbook

        Args:
            scenario_types: List of scenarios to compare (default: all presets)

        Returns:
            BytesIO object containing the Excel file
        """
        # Start with single scenario export
        wb = Workbook()

        # Create all sheets from single scenario
        self._create_cover_sheet(wb)
        self._create_assumptions_sheet(wb)
        self._create_dcf_sheet(wb, 'uss')
        self._create_dcf_sheet(wb, 'nippon')
        self._create_fcf_reconciliation_sheet(wb)
        self._create_consolidated_sheet(wb)

        for segment in Segment:
            self._create_segment_sheet(wb, segment.value)

        self._create_wacc_sheet(wb)
        self._create_projects_sheet(wb)
        self._create_equity_bridge_sheet(wb)
        self._create_financing_sheet(wb)

        # Add synergy sheet if synergies are enabled
        if self.analysis.get('synergy_schedule') is not None:
            self._create_synergy_sheet(wb)

        # Add multi-scenario sheets
        self._create_scenario_comparison_sheet(wb, scenario_types)
        self._create_probability_weighted_sheet(wb)

        # Remove default empty sheet
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_cover_sheet(self, wb: Workbook):
        """Create cover sheet with title, timestamp, and key metrics"""
        ws = wb.active
        ws.title = "Cover"

        # Title
        ws['A1'] = "USS / Nippon Steel Financial Model"
        ws['A1'].font = self.styler.title_font
        ws['A2'] = "DCF Valuation Analysis"
        ws['A2'].font = self.styler.subtitle_font

        # Metadata
        ws['A4'] = "Generated:"
        ws['B4'] = self.timestamp
        ws['A5'] = "Scenario:"
        ws['B5'] = self.scenario.name
        ws['A6'] = "Execution Factor:"
        ws['B6'] = f"{self.execution_factor:.0%}"

        # Key Metrics Summary
        ws['A8'] = "KEY METRICS SUMMARY"
        ws['A8'].font = self.styler.section_font
        ws['A8'].fill = self.styler.section_fill
        ws.merge_cells('A8:D8')

        val_uss = self.analysis['val_uss']
        val_nippon = self.analysis['val_nippon']
        consolidated = self.analysis['consolidated']

        metrics = [
            ("USS Standalone Value", f"${val_uss['share_price']:.2f}/share"),
            ("Nippon View Value", f"${val_nippon['share_price']:.2f}/share"),
            ("Nippon Offer Price", "$55.00/share"),
            ("vs Offer (USS)", f"${55 - val_uss['share_price']:+.2f}"),
            ("vs Offer (Nippon)", f"${val_nippon['share_price'] - 55:+.2f}"),
            ("", ""),
            ("USS WACC", f"{self.scenario.uss_wacc:.1%}"),
            ("Nippon IRP-Adjusted WACC", f"{self.analysis['usd_wacc']:.2%}"),
            ("WACC Advantage", f"{(self.scenario.uss_wacc - self.analysis['usd_wacc'])*100:.0f} bps"),
            ("", ""),
            ("10-Year FCF", f"${consolidated['FCF'].sum():,.0f}M"),
            ("Terminal EBITDA", f"${consolidated['Total_EBITDA'].iloc[-1]:,.0f}M"),
            ("Avg EBITDA Margin", f"{consolidated['EBITDA_Margin'].mean():.1%}"),
            ("", ""),
            ("Enterprise Value (Blended)", f"${val_uss['ev_blended']:,.0f}M"),
            ("Terminal Value (Gordon)", f"${val_uss['tv_gordon']:,.0f}M"),
            ("Terminal Value (Exit)", f"${val_uss['tv_exit']:,.0f}M"),
        ]

        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=9, column=col, value=header)
        self.styler.style_header_row(ws, 9, 1, 2)

        for i, (metric, value) in enumerate(metrics):
            ws.cell(row=10 + i, column=1, value=metric)
            ws.cell(row=10 + i, column=2, value=value)
            if metric:
                self.styler.style_data_cell(ws, 10 + i, 1)
                self.styler.style_data_cell(ws, 10 + i, 2)

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_assumptions_sheet(self, wb: Workbook):
        """Create sheet with all model assumptions"""
        ws = wb.create_sheet("Key Assumptions")

        # Title
        ws['A1'] = "MODEL ASSUMPTIONS"
        ws['A1'].font = self.styler.title_font

        row = 3

        # Price Scenario
        self.styler.style_section_header(ws, row, "STEEL PRICE SCENARIO", 1, 3)
        row += 1

        price_assumptions = [
            ("Scenario Name", self.scenario.price_scenario.name),
            ("US HRC Factor", f"{self.scenario.price_scenario.hrc_us_factor:.2f}x"),
            ("US CRC Factor", f"{self.scenario.price_scenario.crc_us_factor:.2f}x"),
            ("Coated Factor", f"{self.scenario.price_scenario.coated_us_factor:.2f}x"),
            ("EU HRC Factor", f"{self.scenario.price_scenario.hrc_eu_factor:.2f}x"),
            ("OCTG Factor", f"{self.scenario.price_scenario.octg_factor:.2f}x"),
            ("Annual Price Growth", f"{self.scenario.price_scenario.annual_price_growth:.1%}"),
        ]

        for param, value in price_assumptions:
            ws.cell(row=row, column=1, value=param)
            ws.cell(row=row, column=2, value=value)
            self.styler.style_data_cell(ws, row, 1)
            self.styler.style_data_cell(ws, row, 2)
            row += 1

        row += 1

        # Benchmark Prices
        self.styler.style_section_header(ws, row, "BENCHMARK PRICES (2023 Base)", 1, 3)
        row += 1

        for benchmark, price in self.custom_benchmarks.items():
            ws.cell(row=row, column=1, value=benchmark.upper().replace('_', ' '))
            ws.cell(row=row, column=2, value=price)
            ws.cell(row=row, column=2).number_format = self.styler.currency_format
            self.styler.style_data_cell(ws, row, 1)
            self.styler.style_data_cell(ws, row, 2)
            row += 1

        row += 1

        # Volume Scenario
        self.styler.style_section_header(ws, row, "VOLUME SCENARIO", 1, 3)
        row += 1

        vol = self.scenario.volume_scenario
        volume_assumptions = [
            ("Scenario Name", vol.name),
            ("Flat-Rolled Volume Factor", f"{vol.flat_rolled_volume_factor:.2f}x"),
            ("Mini Mill Volume Factor", f"{vol.mini_mill_volume_factor:.2f}x"),
            ("USSE Volume Factor", f"{vol.usse_volume_factor:.2f}x"),
            ("Tubular Volume Factor", f"{vol.tubular_volume_factor:.2f}x"),
            ("Flat-Rolled Growth Adj", f"{vol.flat_rolled_growth_adj:.1%}"),
            ("Mini Mill Growth Adj", f"{vol.mini_mill_growth_adj:.1%}"),
            ("USSE Growth Adj", f"{vol.usse_growth_adj:.1%}"),
            ("Tubular Growth Adj", f"{vol.tubular_growth_adj:.1%}"),
        ]

        for param, value in volume_assumptions:
            ws.cell(row=row, column=1, value=param)
            ws.cell(row=row, column=2, value=value)
            self.styler.style_data_cell(ws, row, 1)
            self.styler.style_data_cell(ws, row, 2)
            row += 1

        row += 1

        # WACC Parameters
        self.styler.style_section_header(ws, row, "WACC & VALUATION PARAMETERS", 1, 3)
        row += 1

        wacc_assumptions = [
            ("USS WACC", f"{self.scenario.uss_wacc:.1%}"),
            ("Terminal Growth Rate", f"{self.scenario.terminal_growth:.2%}"),
            ("Exit EBITDA Multiple", f"{self.scenario.exit_multiple:.1f}x"),
            ("US 10-Year Treasury", f"{self.scenario.us_10yr:.2%}"),
            ("Japan 10-Year JGB", f"{self.scenario.japan_10yr:.2%}"),
            ("Nippon Equity Risk Premium", f"{self.scenario.nippon_equity_risk_premium:.2%}"),
            ("Nippon Credit Spread", f"{self.scenario.nippon_credit_spread:.2%}"),
            ("Nippon Debt Ratio", f"{self.scenario.nippon_debt_ratio:.0%}"),
            ("Nippon Tax Rate", f"{self.scenario.nippon_tax_rate:.0%}"),
        ]

        if self.scenario.override_irp:
            wacc_assumptions.append(("IRP Override", "Yes"))
            wacc_assumptions.append(("Manual USD WACC", f"{self.scenario.manual_nippon_usd_wacc:.2%}"))

        for param, value in wacc_assumptions:
            ws.cell(row=row, column=1, value=param)
            ws.cell(row=row, column=2, value=value)
            self.styler.style_data_cell(ws, row, 1)
            self.styler.style_data_cell(ws, row, 2)
            row += 1

        row += 1

        # Capital Projects
        self.styler.style_section_header(ws, row, "ENABLED CAPITAL PROJECTS", 1, 3)
        row += 1

        if self.scenario.include_projects:
            projects = get_capital_projects()
            for proj_name in self.scenario.include_projects:
                proj = projects.get(proj_name)
                if proj:
                    total_capex = sum(proj.capex_schedule.values())
                    total_ebitda = sum(proj.ebitda_schedule.values())
                    ws.cell(row=row, column=1, value=proj_name)
                    ws.cell(row=row, column=2, value=f"${total_capex:,.0f}M CapEx")
                    ws.cell(row=row, column=3, value=f"${total_ebitda:,.0f}M EBITDA")
                    self.styler.style_data_cell(ws, row, 1)
                    self.styler.style_data_cell(ws, row, 2)
                    self.styler.style_data_cell(ws, row, 3)
                    row += 1
        else:
            ws.cell(row=row, column=1, value="No capital projects enabled")
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20

    def _create_dcf_sheet(self, wb: Workbook, perspective: str):
        """Create DCF valuation sheet

        Args:
            perspective: 'uss' or 'nippon'
        """
        if perspective == 'uss':
            ws = wb.create_sheet("DCF Valuation - USS")
            val = self.analysis['val_uss']
            wacc = self.scenario.uss_wacc
            title = "USS STANDALONE DCF VALUATION"
        else:
            ws = wb.create_sheet("DCF Valuation - Nippon")
            val = self.analysis['val_nippon']
            wacc = self.analysis['usd_wacc']
            title = "NIPPON VIEW DCF VALUATION (IRP-Adjusted)"

        consolidated = self.analysis['consolidated']
        years = consolidated['Year'].tolist()

        # Title
        ws['A1'] = title
        ws['A1'].font = self.styler.title_font
        ws['A2'] = f"Discount Rate: {wacc:.2%}"
        ws['A2'].font = self.styler.subtitle_font

        row = 4

        # FCF Projection
        self.styler.style_section_header(ws, row, "FREE CASH FLOW PROJECTION", 1, len(years) + 1)
        row += 1

        # Headers
        headers = ["Item"] + [str(y) for y in years]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # FCF row
        ws.cell(row=row, column=1, value="Free Cash Flow")
        for col, fcf in enumerate(val['fcf_list'], 2):
            cell = ws.cell(row=row, column=col, value=fcf)
            cell.number_format = self.styler.currency_format
            value_type = 'positive' if fcf > 0 else 'negative' if fcf < 0 else None
            self.styler.style_data_cell(ws, row, col, value_type)
        self.styler.style_data_cell(ws, row, 1)
        row += 1

        # Discount factors
        ws.cell(row=row, column=1, value="Discount Factor")
        discount_factors = [(1 / (1 + wacc)) ** (i + 1) for i in range(len(years))]
        for col, df in enumerate(discount_factors, 2):
            cell = ws.cell(row=row, column=col, value=df)
            cell.number_format = '0.0000'
            self.styler.style_data_cell(ws, row, col)
        self.styler.style_data_cell(ws, row, 1)
        row += 1

        # PV FCF
        ws.cell(row=row, column=1, value="Present Value FCF")
        for col, pv in enumerate(val['pv_fcf'], 2):
            cell = ws.cell(row=row, column=col, value=pv)
            cell.number_format = self.styler.currency_format
            value_type = 'positive' if pv > 0 else 'negative' if pv < 0 else None
            self.styler.style_data_cell(ws, row, col, value_type)
        self.styler.style_data_cell(ws, row, 1)
        row += 2

        # Valuation Summary
        self.styler.style_section_header(ws, row, "VALUATION SUMMARY", 1, 3)
        row += 1

        summary_items = [
            ("Sum of PV FCF", val['sum_pv_fcf'], self.styler.currency_format),
            ("", "", ""),
            ("Terminal Value - Gordon Growth", "", ""),
            ("  Terminal FCF (Y+1)", val['fcf_list'][-1] * (1 + self.scenario.terminal_growth), self.styler.currency_format),
            ("  Terminal Growth Rate", self.scenario.terminal_growth, self.styler.percent_decimal_format),
            ("  Terminal Value (undiscounted)", val['tv_gordon'], self.styler.currency_format),
            ("  PV of Terminal Value", val['pv_tv_gordon'], self.styler.currency_format),
            ("", "", ""),
            ("Terminal Value - Exit Multiple", "", ""),
            ("  Terminal EBITDA", val['terminal_ebitda'], self.styler.currency_format),
            ("  Exit Multiple", self.scenario.exit_multiple, self.styler.multiple_format),
            ("  Terminal Value (undiscounted)", val['tv_exit'], self.styler.currency_format),
            ("  PV of Terminal Value", val['pv_tv_exit'], self.styler.currency_format),
            ("", "", ""),
            ("Enterprise Value - Gordon", val['ev_gordon'], self.styler.currency_format),
            ("Enterprise Value - Exit Multiple", val['ev_exit'], self.styler.currency_format),
            ("Enterprise Value - Blended", val['ev_blended'], self.styler.currency_format),
            ("", "", ""),
            ("Equity Bridge", val['equity_bridge'], self.styler.currency_format),
            ("Shares Outstanding (M)", val['shares_used'], self.styler.decimal_format),
            ("", "", ""),
            ("Share Price", val['share_price'], self.styler.currency_decimal_format),
        ]

        for item, value, fmt in summary_items:
            ws.cell(row=row, column=1, value=item)
            if value != "":
                cell = ws.cell(row=row, column=2, value=value)
                cell.number_format = fmt
            if item and item != "":
                self.styler.style_data_cell(ws, row, 1)
                if value != "":
                    value_type = 'positive' if isinstance(value, (int, float)) and value > 0 else None
                    if item == "Share Price":
                        value_type = 'positive'
                    self.styler.style_data_cell(ws, row, 2, value_type)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        for col in range(3, len(years) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_fcf_reconciliation_sheet(self, wb: Workbook):
        """Create FCF reconciliation sheet showing EBITDA to FCF bridge"""
        ws = wb.create_sheet("FCF Reconciliation")

        consolidated = self.analysis['consolidated']
        years = consolidated['Year'].tolist()

        # Title
        ws['A1'] = "FREE CASH FLOW RECONCILIATION"
        ws['A1'].font = self.styler.title_font

        row = 3

        # Headers
        headers = ["Item"] + [str(y) for y in years] + ["Total"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # FCF Bridge items
        bridge_items = [
            ("Revenue", "Revenue"),
            ("Total EBITDA", "Total_EBITDA"),
            ("(-) D&A", "DA", -1),
            ("= EBIT", None, None, lambda df: df['Total_EBITDA'] - df['DA']),
            ("(-) Cash Taxes (16.9%)", None, None, lambda df: (df['Total_EBITDA'] - df['DA']) * 0.169),
            ("= NOPAT", "NOPAT"),
            ("(+) D&A", "DA"),
            ("= Gross Cash Flow", "Gross_CF"),
            ("(-) CapEx", "Total_CapEx", -1),
            ("(+/-) Change in NWC", "Delta_WC"),
            ("= Free Cash Flow", "FCF"),
        ]

        for item_name, col_name, *extra in bridge_items:
            ws.cell(row=row, column=1, value=item_name)
            self.styler.style_data_cell(ws, row, 1)

            multiplier = extra[0] if extra and extra[0] is not None else 1
            calc_func = extra[1] if len(extra) > 1 else None

            total = 0
            for col_idx, year in enumerate(years, 2):
                year_data = consolidated[consolidated['Year'] == year]
                if col_name:
                    value = year_data[col_name].values[0] * multiplier
                elif calc_func:
                    value = calc_func(year_data).values[0] * multiplier
                else:
                    value = 0

                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.number_format = self.styler.currency_format
                value_type = 'positive' if value > 0 else 'negative' if value < 0 else None
                self.styler.style_data_cell(ws, row, col_idx, value_type)
                total += value

            # Total column
            cell = ws.cell(row=row, column=len(years) + 2, value=total)
            cell.number_format = self.styler.currency_format
            value_type = 'positive' if total > 0 else 'negative' if total < 0 else None
            self.styler.style_data_cell(ws, row, len(years) + 2, value_type)

            row += 1

        # EBITDA Margin row
        row += 1
        ws.cell(row=row, column=1, value="EBITDA Margin")
        self.styler.style_data_cell(ws, row, 1, 'derived')
        for col_idx, year in enumerate(years, 2):
            year_data = consolidated[consolidated['Year'] == year]
            margin = year_data['EBITDA_Margin'].values[0]
            cell = ws.cell(row=row, column=col_idx, value=margin)
            cell.number_format = self.styler.percent_format
            self.styler.style_data_cell(ws, row, col_idx, 'derived')

        # Column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(years) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_consolidated_sheet(self, wb: Workbook):
        """Create consolidated financial projections sheet"""
        ws = wb.create_sheet("Consolidated Financials")

        consolidated = self.analysis['consolidated']

        # Title
        ws['A1'] = "CONSOLIDATED FINANCIAL PROJECTIONS"
        ws['A1'].font = self.styler.title_font

        row = 3

        # Convert DataFrame to rows
        headers = list(consolidated.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        for _, data_row in consolidated.iterrows():
            for col, header in enumerate(headers, 1):
                value = data_row[header]
                cell = ws.cell(row=row, column=col, value=value)

                # Apply number format based on column
                if header in ['EBITDA_Margin']:
                    cell.number_format = self.styler.percent_format
                elif header in ['Avg_Price_per_ton']:
                    cell.number_format = self.styler.currency_format
                elif header not in ['Year']:
                    cell.number_format = self.styler.number_format

                self.styler.style_data_cell(ws, row, col)
            row += 1

        self.styler.auto_column_width(ws)

    def _create_segment_sheet(self, wb: Workbook, segment_name: str):
        """Create sheet for a specific segment"""
        ws = wb.create_sheet(f"Segment - {segment_name}")

        segment_df = self.analysis['segment_dfs'].get(segment_name)
        if segment_df is None:
            ws['A1'] = f"No data for segment: {segment_name}"
            return

        # Title
        ws['A1'] = f"{segment_name.upper()} SEGMENT PROJECTIONS"
        ws['A1'].font = self.styler.title_font

        row = 3

        # Convert DataFrame to rows
        headers = list(segment_df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        for _, data_row in segment_df.iterrows():
            for col, header in enumerate(headers, 1):
                value = data_row[header]
                cell = ws.cell(row=row, column=col, value=value)

                # Apply number format based on column
                if header in ['EBITDA_Margin']:
                    cell.number_format = self.styler.percent_format
                elif header in ['Price_per_ton']:
                    cell.number_format = self.styler.currency_format
                elif header not in ['Year', 'Segment']:
                    cell.number_format = self.styler.number_format

                self.styler.style_data_cell(ws, row, col)
            row += 1

        self.styler.auto_column_width(ws)

    def _create_wacc_sheet(self, wb: Workbook):
        """Create WACC analysis sheet"""
        ws = wb.create_sheet("WACC Analysis")

        # Title
        ws['A1'] = "WACC ANALYSIS"
        ws['A1'].font = self.styler.title_font

        row = 3

        # USS WACC
        self.styler.style_section_header(ws, row, "USS STANDALONE WACC", 1, 3)
        row += 1

        ws.cell(row=row, column=1, value="USS WACC (Input)")
        ws.cell(row=row, column=2, value=self.scenario.uss_wacc)
        ws.cell(row=row, column=2).number_format = self.styler.percent_format
        self.styler.style_data_cell(ws, row, 1)
        self.styler.style_data_cell(ws, row, 2)
        row += 2

        # Nippon WACC Calculation
        self.styler.style_section_header(ws, row, "NIPPON USD WACC CALCULATION (IRP Method)", 1, 3)
        row += 1

        jpy_wacc = self.analysis['jpy_wacc']
        usd_wacc = self.analysis['usd_wacc']

        # Calculate intermediate values
        nippon_cost_of_equity = self.scenario.japan_10yr + self.scenario.nippon_equity_risk_premium
        nippon_cost_of_debt = self.scenario.japan_10yr + self.scenario.nippon_credit_spread
        equity_weight = 1 - self.scenario.nippon_debt_ratio
        debt_weight = self.scenario.nippon_debt_ratio

        wacc_items = [
            ("Step 1: Nippon Cost of Capital in JPY", "", ""),
            ("  Japan 10-Year JGB", self.scenario.japan_10yr, self.styler.percent_decimal_format),
            ("  + Equity Risk Premium", self.scenario.nippon_equity_risk_premium, self.styler.percent_decimal_format),
            ("  = Cost of Equity (JPY)", nippon_cost_of_equity, self.styler.percent_decimal_format),
            ("", "", ""),
            ("  Japan 10-Year JGB", self.scenario.japan_10yr, self.styler.percent_decimal_format),
            ("  + Credit Spread", self.scenario.nippon_credit_spread, self.styler.percent_decimal_format),
            ("  = Cost of Debt (JPY)", nippon_cost_of_debt, self.styler.percent_decimal_format),
            ("", "", ""),
            ("  Equity Weight", equity_weight, self.styler.percent_format),
            ("  Debt Weight", debt_weight, self.styler.percent_format),
            ("  Tax Rate", self.scenario.nippon_tax_rate, self.styler.percent_format),
            ("", "", ""),
            ("  = JPY WACC", jpy_wacc, self.styler.percent_decimal_format),
            ("", "", ""),
            ("Step 2: Convert to USD via IRP", "", ""),
            ("  Formula: (1 + JPY WACC) x (1 + US 10Y) / (1 + JP 10Y) - 1", "", ""),
            ("  US 10-Year Treasury", self.scenario.us_10yr, self.styler.percent_decimal_format),
            ("  Japan 10-Year JGB", self.scenario.japan_10yr, self.styler.percent_decimal_format),
            ("", "", ""),
        ]

        if self.scenario.override_irp:
            wacc_items.extend([
                ("  IRP-Implied USD WACC", (1 + jpy_wacc) * (1 + self.scenario.us_10yr) / (1 + self.scenario.japan_10yr) - 1, self.styler.percent_decimal_format),
                ("  OVERRIDE: Manual USD WACC", self.scenario.manual_nippon_usd_wacc, self.styler.percent_decimal_format),
                ("  = Nippon USD WACC (Used)", usd_wacc, self.styler.percent_decimal_format),
            ])
        else:
            wacc_items.append(("  = Nippon USD WACC", usd_wacc, self.styler.percent_decimal_format))

        wacc_items.extend([
            ("", "", ""),
            ("WACC Comparison", "", ""),
            ("  USS WACC", self.scenario.uss_wacc, self.styler.percent_decimal_format),
            ("  Nippon USD WACC", usd_wacc, self.styler.percent_decimal_format),
            ("  WACC Advantage (bps)", (self.scenario.uss_wacc - usd_wacc) * 10000, "#,##0"),
        ])

        for item, value, fmt in wacc_items:
            ws.cell(row=row, column=1, value=item)
            if value != "":
                cell = ws.cell(row=row, column=2, value=value)
                if fmt:
                    cell.number_format = fmt
            if item and item != "":
                self.styler.style_data_cell(ws, row, 1)
                if value != "":
                    self.styler.style_data_cell(ws, row, 2)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_projects_sheet(self, wb: Workbook):
        """Create capital projects sheet"""
        ws = wb.create_sheet("Capital Projects")

        # Title
        ws['A1'] = "CAPITAL PROJECTS ANALYSIS"
        ws['A1'].font = self.styler.title_font

        projects = get_capital_projects()
        years = list(range(2024, 2034))

        row = 3

        # CapEx Schedule
        self.styler.style_section_header(ws, row, "CAPEX SCHEDULE ($M)", 1, len(years) + 2)
        row += 1

        headers = ["Project"] + [str(y) for y in years] + ["Total"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        for proj_name, proj in projects.items():
            enabled = proj_name in self.scenario.include_projects
            ws.cell(row=row, column=1, value=f"{proj_name}{' (Enabled)' if enabled else ''}")
            self.styler.style_data_cell(ws, row, 1, 'positive' if enabled else None)

            total = 0
            for col_idx, year in enumerate(years, 2):
                value = proj.capex_schedule.get(year, 0)
                if enabled:
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.number_format = self.styler.currency_format
                else:
                    cell = ws.cell(row=row, column=col_idx, value=0)
                self.styler.style_data_cell(ws, row, col_idx, 'positive' if enabled and value > 0 else None)
                if enabled:
                    total += value

            cell = ws.cell(row=row, column=len(years) + 2, value=total)
            cell.number_format = self.styler.currency_format
            self.styler.style_data_cell(ws, row, len(years) + 2, 'positive' if enabled else None)
            row += 1

        row += 2

        # EBITDA Schedule
        self.styler.style_section_header(ws, row, "INCREMENTAL EBITDA SCHEDULE ($M)", 1, len(years) + 2)
        row += 1

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        for proj_name, proj in projects.items():
            enabled = proj_name in self.scenario.include_projects
            ws.cell(row=row, column=1, value=f"{proj_name}{' (Enabled)' if enabled else ''}")
            self.styler.style_data_cell(ws, row, 1, 'positive' if enabled else None)

            total = 0
            for col_idx, year in enumerate(years, 2):
                value = proj.ebitda_schedule.get(year, 0)
                # Apply execution factor for non-BR2 projects
                if enabled and proj_name != 'BR2 Mini Mill':
                    value *= self.execution_factor
                elif not enabled:
                    value = 0

                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.number_format = self.styler.currency_format
                self.styler.style_data_cell(ws, row, col_idx, 'positive' if enabled and value > 0 else None)
                if enabled:
                    total += value

            cell = ws.cell(row=row, column=len(years) + 2, value=total)
            cell.number_format = self.styler.currency_format
            self.styler.style_data_cell(ws, row, len(years) + 2, 'positive' if enabled else None)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, len(years) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 10

    def _create_equity_bridge_sheet(self, wb: Workbook):
        """Create equity bridge sheet"""
        ws = wb.create_sheet("Equity Bridge")

        # Title
        ws['A1'] = "EQUITY BRIDGE: ENTERPRISE VALUE TO EQUITY VALUE"
        ws['A1'].font = self.styler.title_font

        val_uss = self.analysis['val_uss']
        val_nippon = self.analysis['val_nippon']
        financing = self.analysis.get('financing_impact', {})

        row = 3

        # Headers
        headers = ["Item", "USS Standalone", "Nippon View"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # Bridge items
        base_debt = 4222.0
        uss_total_debt = financing.get('total_debt', base_debt)
        pension = 126.0
        leases = 117.0
        cash = 3013.9
        investments = 761.0
        uss_shares = financing.get('total_shares', 225.0)
        nippon_shares = 225.0

        bridge_items = [
            ("Enterprise Value (Blended)", val_uss['ev_blended'], val_nippon['ev_blended']),
            ("", "", ""),
            ("Less: Total Debt", -uss_total_debt, -base_debt),
            ("Less: Pension Obligations", -pension, -pension),
            ("Less: Operating Leases", -leases, -leases),
            ("Plus: Cash & Equivalents", cash, cash),
            ("Plus: Equity Investments", investments, investments),
            ("", "", ""),
            ("= Equity Value", val_uss['ev_blended'] + val_uss['equity_bridge'], val_nippon['ev_blended'] + val_nippon['equity_bridge']),
            ("", "", ""),
            ("Shares Outstanding (M)", uss_shares, nippon_shares),
            ("", "", ""),
            ("= Share Price", val_uss['share_price'], val_nippon['share_price']),
        ]

        for item, uss_val, nippon_val in bridge_items:
            ws.cell(row=row, column=1, value=item)
            if uss_val != "":
                cell = ws.cell(row=row, column=2, value=uss_val)
                if "Share Price" in item:
                    cell.number_format = self.styler.currency_decimal_format
                elif "Shares" in item:
                    cell.number_format = self.styler.decimal_format
                else:
                    cell.number_format = self.styler.currency_format
            if nippon_val != "":
                cell = ws.cell(row=row, column=3, value=nippon_val)
                if "Share Price" in item:
                    cell.number_format = self.styler.currency_decimal_format
                elif "Shares" in item:
                    cell.number_format = self.styler.decimal_format
                else:
                    cell.number_format = self.styler.currency_format

            if item:
                self.styler.style_data_cell(ws, row, 1)
                self.styler.style_data_cell(ws, row, 2)
                self.styler.style_data_cell(ws, row, 3)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

    def _create_financing_sheet(self, wb: Workbook):
        """Create financing impact sheet for USS standalone"""
        ws = wb.create_sheet("Financing Impact")

        # Title
        ws['A1'] = "USS STANDALONE FINANCING IMPACT"
        ws['A1'].font = self.styler.title_font
        ws['A2'] = "Impact of financing $14B capital program without Nippon's balance sheet"
        ws['A2'].font = Font(italic=True)

        financing = self.analysis.get('financing_impact', {})

        row = 4

        if not financing or financing.get('financing_gap', 0) == 0:
            ws.cell(row=row, column=1, value="No significant financing impact (FCF covers capital needs)")
            return

        # Financing summary
        self.styler.style_section_header(ws, row, "FINANCING REQUIREMENTS", 1, 2)
        row += 1

        items = [
            ("Incremental Project CapEx", financing.get('incremental_capex', 0), self.styler.currency_format),
            ("Financing Gap (Cumulative Negative FCF)", financing.get('financing_gap', 0), self.styler.currency_format),
            ("", "", ""),
            ("New Debt (50%)", financing.get('new_debt', 0), self.styler.currency_format),
            ("New Equity (50%)", financing.get('new_equity', 0), self.styler.currency_format),
            ("", "", ""),
            ("Debt Impact", "", ""),
            ("  Total Debt (After)", financing.get('total_debt', 0), self.styler.currency_format),
            ("  Debt / EBITDA", financing.get('debt_to_ebitda', 0), self.styler.multiple_format),
            ("  Annual Interest Expense", financing.get('annual_interest_expense', 0), self.styler.currency_format),
            ("  WACC Adjustment", financing.get('wacc_adjustment', 0), self.styler.percent_decimal_format),
            ("  Adjusted WACC", financing.get('adjusted_wacc', 0), self.styler.percent_decimal_format),
            ("", "", ""),
            ("Equity Impact", "", ""),
            ("  Current Shares (M)", 225.0, self.styler.decimal_format),
            ("  New Shares Issued (M)", financing.get('new_shares', 0), self.styler.decimal_format),
            ("  Total Shares (M)", financing.get('total_shares', 0), self.styler.decimal_format),
            ("  Dilution %", financing.get('dilution_pct', 0), self.styler.percent_format),
        ]

        for item, value, fmt in items:
            ws.cell(row=row, column=1, value=item)
            if value != "":
                cell = ws.cell(row=row, column=2, value=value)
                cell.number_format = fmt
            if item and item != "":
                self.styler.style_data_cell(ws, row, 1)
                if value != "":
                    self.styler.style_data_cell(ws, row, 2)
            row += 1

        row += 2

        # Key insight
        ws.cell(row=row, column=1, value="KEY INSIGHT:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="USS cannot fund $14B in projects without dilutive financing.")
        ws.cell(row=row + 1, column=1, value="Nippon's balance sheet absorbs this cost, creating value for USS shareholders.")

        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _create_synergy_sheet(self, wb: Workbook):
        """Create synergy analysis sheet"""
        ws = wb.create_sheet("Synergy Analysis")

        synergy_schedule = self.analysis.get('synergy_schedule')
        synergy_value = self.analysis.get('synergy_value')
        synergies = self.scenario.synergies

        if synergy_schedule is None or synergies is None:
            ws['A1'] = "No synergies configured for this scenario"
            return

        # Title
        ws['A1'] = "SYNERGY & TECHNOLOGY TRANSFER ANALYSIS"
        ws['A1'].font = self.styler.title_font
        ws['A2'] = f"Synergy Preset: {synergies.name}"
        ws['A2'].font = self.styler.subtitle_font

        row = 4

        # Summary section
        self.styler.style_section_header(ws, row, "SYNERGY VALUE SUMMARY", 1, 2)
        row += 1

        if synergy_value:
            summary_items = [
                ("NPV of Synergies", synergy_value['npv_synergies'], self.styler.currency_format),
                ("Synergy Value per Share", synergy_value['synergy_value_per_share'], self.styler.currency_decimal_format),
                ("Run-Rate Synergies (2033)", synergy_value['run_rate_synergies'], self.styler.currency_format),
                ("Total Integration Costs", synergy_value['total_integration_costs'], self.styler.currency_format),
                ("", "", ""),
                ("Operating Run-Rate", synergies.operating.get_total_run_rate(), self.styler.currency_format),
                ("Revenue Synergy EBITDA", synergies.revenue.get_run_rate_ebitda(), self.styler.currency_format),
                ("Integration Total", synergies.integration.get_total_cost(), self.styler.currency_format),
            ]

            for item, value, fmt in summary_items:
                ws.cell(row=row, column=1, value=item)
                if value != "":
                    cell = ws.cell(row=row, column=2, value=value)
                    cell.number_format = fmt
                if item:
                    self.styler.style_data_cell(ws, row, 1)
                    if value != "":
                        value_type = 'positive' if isinstance(value, (int, float)) and value > 0 else None
                        self.styler.style_data_cell(ws, row, 2, value_type)
                row += 1

        row += 1

        # Operating Synergies Detail
        self.styler.style_section_header(ws, row, "OPERATING SYNERGIES ASSUMPTIONS", 1, 3)
        row += 1

        op = synergies.operating
        op_items = [
            ("Procurement Savings (Annual)", op.procurement_savings_annual, op.procurement_confidence),
            ("Logistics Savings (Annual)", op.logistics_savings_annual, op.logistics_confidence),
            ("Overhead Savings (Annual)", op.overhead_savings_annual, op.overhead_confidence),
        ]

        headers = ["Category", "Run-Rate ($M)", "Confidence"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, 3)
        row += 1

        for item, value, conf in op_items:
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = self.styler.currency_format
            ws.cell(row=row, column=3, value=conf)
            ws.cell(row=row, column=3).number_format = self.styler.percent_format
            for col in range(1, 4):
                self.styler.style_data_cell(ws, row, col)
            row += 1

        row += 1

        # Technology Transfer Detail
        self.styler.style_section_header(ws, row, "TECHNOLOGY TRANSFER ASSUMPTIONS", 1, 3)
        row += 1

        tech = synergies.technology
        tech_items = [
            ("Yield Improvement", tech.yield_improvement_pct, self.styler.percent_format),
            ("Quality Price Premium", tech.quality_price_premium_pct, self.styler.percent_format),
            ("Conversion Cost Reduction", tech.conversion_cost_reduction_pct, self.styler.percent_format),
            ("Confidence Factor", tech.confidence, self.styler.percent_format),
        ]

        for item, value, fmt in tech_items:
            ws.cell(row=row, column=1, value=item)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = fmt
            self.styler.style_data_cell(ws, row, 1)
            self.styler.style_data_cell(ws, row, 2)
            row += 1

        row += 1

        # Revenue Synergies Detail
        self.styler.style_section_header(ws, row, "REVENUE SYNERGIES ASSUMPTIONS", 1, 3)
        row += 1

        rev = synergies.revenue
        rev_items = [
            ("Cross-Sell Revenue (Annual)", rev.cross_sell_revenue_annual, rev.cross_sell_margin, rev.cross_sell_confidence),
            ("Product Mix Uplift (Annual)", rev.product_mix_revenue_uplift, rev.product_mix_margin, rev.product_mix_confidence),
        ]

        headers = ["Category", "Revenue ($M)", "Margin", "Confidence"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, 4)
        row += 1

        for item, rev_val, margin, conf in rev_items:
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=rev_val)
            ws.cell(row=row, column=2).number_format = self.styler.currency_format
            ws.cell(row=row, column=3, value=margin)
            ws.cell(row=row, column=3).number_format = self.styler.percent_format
            ws.cell(row=row, column=4, value=conf)
            ws.cell(row=row, column=4).number_format = self.styler.percent_format
            for col in range(1, 5):
                self.styler.style_data_cell(ws, row, col)
            row += 1

        row += 2

        # Year-by-year schedule
        self.styler.style_section_header(ws, row, "SYNERGY SCHEDULE ($M)", 1, len(synergy_schedule) + 1)
        row += 1

        # Headers
        headers = list(synergy_schedule.columns)
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # Data rows
        for _, data_row in synergy_schedule.iterrows():
            for col, header in enumerate(headers, 1):
                value = data_row[header]
                cell = ws.cell(row=row, column=col, value=value)

                if header != 'Year':
                    cell.number_format = self.styler.currency_format

                value_type = None
                if header == 'Integration_Cost' and value > 0:
                    value_type = 'negative'
                elif header in ['Total_Synergy_EBITDA', 'Cumulative_Synergy'] and value > 0:
                    value_type = 'positive'

                self.styler.style_data_cell(ws, row, col, value_type)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        for col in range(5, len(headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_scenario_comparison_sheet(self, wb: Workbook, scenario_types: List[ScenarioType] = None):
        """Create scenario comparison sheet"""
        ws = wb.create_sheet("Scenario Comparison")

        # Title
        ws['A1'] = "SCENARIO COMPARISON"
        ws['A1'].font = self.styler.title_font

        # Run comparison
        comparison = compare_scenarios(
            scenario_types=scenario_types,
            execution_factor=self.execution_factor,
            custom_benchmarks=self.custom_benchmarks
        )

        row = 3

        # Headers
        headers = list(comparison.columns)
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # Data rows
        for _, data_row in comparison.iterrows():
            for col, header in enumerate(headers, 1):
                value = data_row[header]
                cell = ws.cell(row=row, column=col, value=value)

                # Apply formatting
                if 'share' in header.lower() or '$' in header:
                    cell.number_format = self.styler.currency_decimal_format
                elif 'margin' in header.lower() or '%' in header:
                    cell.number_format = self.styler.percent_format
                elif 'ebitda' in header.lower() and 'ev' in header.lower():
                    cell.number_format = self.styler.multiple_format

                self.styler.style_data_cell(ws, row, col)
            row += 1

        self.styler.auto_column_width(ws)

    def _create_probability_weighted_sheet(self, wb: Workbook):
        """Create probability-weighted valuation sheet"""
        ws = wb.create_sheet("Probability-Weighted")

        # Title
        ws['A1'] = "PROBABILITY-WEIGHTED EXPECTED VALUE"
        ws['A1'].font = self.styler.title_font

        row = 3

        try:
            pw_results = calculate_probability_weighted_valuation(
                custom_benchmarks=self.custom_benchmarks
            )
        except ValueError as e:
            ws.cell(row=row, column=1, value=f"Error calculating probability-weighted value: {str(e)}")
            return

        # Summary metrics
        self.styler.style_section_header(ws, row, "EXPECTED VALUE SUMMARY", 1, 2)
        row += 1

        summary_items = [
            ("Weighted USS Value ($/share)", pw_results['weighted_uss_value_per_share'], self.styler.currency_decimal_format),
            ("Weighted Nippon Value ($/share)", pw_results['weighted_nippon_value_per_share'], self.styler.currency_decimal_format),
            ("Nippon Offer Price", pw_results['offer_price'], self.styler.currency_decimal_format),
            ("", "", ""),
            ("USS Premium to Offer", pw_results['uss_premium_to_offer'] / 100, self.styler.percent_format),
            ("Nippon Discount to Offer", pw_results['nippon_discount_to_offer'] / 100, self.styler.percent_format),
            ("", "", ""),
            ("Weighted 10Y FCF ($M)", pw_results['weighted_ten_year_fcf'], self.styler.currency_format),
            ("Weighted Avg EBITDA ($M)", pw_results['weighted_avg_ebitda'], self.styler.currency_format),
            ("Total Probability", pw_results['total_probability'], self.styler.percent_format),
        ]

        for item, value, fmt in summary_items:
            ws.cell(row=row, column=1, value=item)
            if value != "":
                cell = ws.cell(row=row, column=2, value=value)
                cell.number_format = fmt
            if item:
                self.styler.style_data_cell(ws, row, 1)
                if value != "":
                    self.styler.style_data_cell(ws, row, 2)
            row += 1

        row += 2

        # Scenario breakdown
        self.styler.style_section_header(ws, row, "SCENARIO BREAKDOWN", 1, 5)
        row += 1

        headers = ["Scenario", "Probability", "USS Value", "Nippon Value", "10Y FCF"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        for scenario_type, result in pw_results['scenario_results'].items():
            ws.cell(row=row, column=1, value=result['name'])
            ws.cell(row=row, column=2, value=result['probability'])
            ws.cell(row=row, column=2).number_format = self.styler.percent_format
            ws.cell(row=row, column=3, value=result['uss_value_per_share'])
            ws.cell(row=row, column=3).number_format = self.styler.currency_decimal_format
            ws.cell(row=row, column=4, value=result['nippon_value_per_share'])
            ws.cell(row=row, column=4).number_format = self.styler.currency_decimal_format
            ws.cell(row=row, column=5, value=result['ten_year_fcf'])
            ws.cell(row=row, column=5).number_format = self.styler.currency_format

            for col in range(1, 6):
                self.styler.style_data_cell(ws, row, col)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15


def get_export_filename(scenario_name: str, export_type: str = "single") -> str:
    """Generate export filename with timestamp

    Args:
        scenario_name: Name of the scenario
        export_type: 'single' or 'multi'

    Returns:
        Filename string
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    scenario_slug = scenario_name.replace(" ", "_").replace("-", "_")[:30]

    if export_type == "multi":
        return f"USS_Model_Comparison_{timestamp}.xlsx"
    elif export_type == "formula":
        return f"USS_Model_Formula_{scenario_slug}_{timestamp}.xlsx"
    else:
        return f"USS_Model_{scenario_slug}_{timestamp}.xlsx"


class FormulaModelExporter:
    """Export financial model to Excel with working formulas

    Creates a fully functional Excel model where changing inputs
    automatically recalculates all outputs through Excel formulas.
    """

    def __init__(self, scenario: ModelScenario, execution_factor: float = 1.0,
                 custom_benchmarks: dict = None):
        self.scenario = scenario
        self.execution_factor = execution_factor
        self.custom_benchmarks = custom_benchmarks or BENCHMARK_PRICES_2023
        self.styler = ExcelStyler()
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Run model to get projection data
        self.model = PriceVolumeModel(scenario, execution_factor, custom_benchmarks)
        self.analysis = self.model.run_full_analysis()

        self.years = list(range(2024, 2034))
        self.num_years = len(self.years)

        # Cell reference tracking - populated during sheet creation
        self.cell_refs = {}

    def export_with_formulas(self) -> BytesIO:
        """Export model with working Excel formulas"""
        wb = Workbook()

        # Create sheets in order
        self._create_inputs_sheet(wb)
        self._create_projections_sheet(wb)
        self._create_dcf_formula_sheet(wb, 'USS')
        self._create_dcf_formula_sheet(wb, 'Nippon')
        self._create_wacc_formula_sheet(wb)
        self._create_equity_bridge_formula_sheet(wb)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_inputs_sheet(self, wb: Workbook):
        """Create inputs sheet with all model assumptions

        Tracks cell positions in self.cell_refs for use by other sheets.
        """
        ws = wb.active
        ws.title = "Inputs"

        # Title
        ws['A1'] = "MODEL INPUTS - Change values here to update all calculations"
        ws['A1'].font = self.styler.title_font
        ws['A2'] = f"Scenario: {self.scenario.name} | Generated: {self.timestamp}"
        ws['A2'].font = Font(italic=True)

        # Highlight input cells
        input_fill = PatternFill(start_color="FFFFD9", end_color="FFFFD9", fill_type="solid")

        def write_input(ws, row, label, key, value, fmt):
            """Write an input row and track its cell reference"""
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=key)
            cell = ws.cell(row=row, column=3, value=value)
            cell.number_format = fmt
            cell.fill = input_fill
            self.styler.style_data_cell(ws, row, 1)
            self.styler.style_data_cell(ws, row, 3)
            # Track cell reference
            self.cell_refs[key] = f"$C${row}"

        # === WACC & VALUATION SECTION ===
        row = 4
        self.styler.style_section_header(ws, row, "WACC & VALUATION PARAMETERS", 1, 4)
        row += 1

        write_input(ws, row, 'USS WACC', 'uss_wacc', self.scenario.uss_wacc, '0.0%')
        row += 1
        write_input(ws, row, 'Terminal Growth Rate', 'terminal_growth', self.scenario.terminal_growth, '0.00%')
        row += 1
        write_input(ws, row, 'Exit EBITDA Multiple', 'exit_multiple', self.scenario.exit_multiple, '0.0"x"')
        row += 1
        write_input(ws, row, 'Cash Tax Rate', 'tax_rate', 0.169, '0.0%')
        row += 1

        # === IRP PARAMETERS SECTION ===
        row += 1
        self.styler.style_section_header(ws, row, "INTEREST RATE PARITY PARAMETERS", 1, 4)
        row += 1

        write_input(ws, row, 'US 10-Year Treasury', 'us_10yr', self.scenario.us_10yr, '0.00%')
        row += 1
        write_input(ws, row, 'Japan 10-Year JGB', 'japan_10yr', self.scenario.japan_10yr, '0.00%')
        row += 1
        write_input(ws, row, 'Nippon Equity Risk Premium', 'nippon_erp', self.scenario.nippon_equity_risk_premium, '0.00%')
        row += 1
        write_input(ws, row, 'Nippon Credit Spread', 'nippon_cs', self.scenario.nippon_credit_spread, '0.00%')
        row += 1
        write_input(ws, row, 'Nippon Debt Ratio', 'nippon_debt_ratio', self.scenario.nippon_debt_ratio, '0%')
        row += 1
        write_input(ws, row, 'Nippon Tax Rate', 'nippon_tax_rate', self.scenario.nippon_tax_rate, '0%')
        row += 1

        # === BALANCE SHEET SECTION ===
        row += 1
        self.styler.style_section_header(ws, row, "BALANCE SHEET ITEMS ($M)", 1, 4)
        row += 1

        write_input(ws, row, 'Total Debt', 'total_debt', 4222.0, '$#,##0')
        row += 1
        write_input(ws, row, 'Pension Obligations', 'pension', 126.0, '$#,##0')
        row += 1
        write_input(ws, row, 'Operating Leases', 'leases', 117.0, '$#,##0')
        row += 1
        write_input(ws, row, 'Cash & Equivalents', 'cash', 3013.9, '$#,##0')
        row += 1
        write_input(ws, row, 'Equity Investments', 'investments', 761.0, '$#,##0')
        row += 1
        write_input(ws, row, 'Shares Outstanding (M)', 'shares', 225.0, '#,##0.0')
        row += 1

        # === CALCULATED VALUES SECTION (with formulas) ===
        row += 1
        self.styler.style_section_header(ws, row, "CALCULATED VALUES (Formulas)", 1, 4)
        row += 1

        # Get tracked cell references
        jgb = self.cell_refs['japan_10yr']
        erp = self.cell_refs['nippon_erp']
        cs = self.cell_refs['nippon_cs']
        dr = self.cell_refs['nippon_debt_ratio']
        tr = self.cell_refs['nippon_tax_rate']
        us10 = self.cell_refs['us_10yr']
        uss_wacc = self.cell_refs['uss_wacc']

        # Cost of Equity = JGB + ERP (no Inputs! prefix needed - same sheet)
        ws.cell(row=row, column=1, value="Nippon Cost of Equity (JPY)")
        ws.cell(row=row, column=3, value=f"={jgb}+{erp}")
        ws.cell(row=row, column=3).number_format = '0.00%'
        self.styler.style_data_cell(ws, row, 1, 'derived')
        self.styler.style_data_cell(ws, row, 3, 'derived')
        row += 1

        # Cost of Debt = JGB + CS
        ws.cell(row=row, column=1, value="Nippon Cost of Debt (JPY)")
        ws.cell(row=row, column=3, value=f"={jgb}+{cs}")
        ws.cell(row=row, column=3).number_format = '0.00%'
        self.styler.style_data_cell(ws, row, 1, 'derived')
        self.styler.style_data_cell(ws, row, 3, 'derived')
        row += 1

        # JPY WACC = (1-D)*Ke + D*Kd*(1-T)
        ws.cell(row=row, column=1, value="Nippon JPY WACC")
        formula = f"=(1-{dr})*({jgb}+{erp})+{dr}*({jgb}+{cs})*(1-{tr})"
        ws.cell(row=row, column=3, value=formula)
        ws.cell(row=row, column=3).number_format = '0.00%'
        self.styler.style_data_cell(ws, row, 1, 'derived')
        self.styler.style_data_cell(ws, row, 3, 'derived')
        jpy_wacc_ref = f"$C${row}"
        self.cell_refs['jpy_wacc'] = jpy_wacc_ref
        row += 1

        # USD WACC via IRP: (1+JPY_WACC)*(1+US_10Y)/(1+JP_10Y)-1
        ws.cell(row=row, column=1, value="Nippon USD WACC (IRP)")
        formula = f"=(1+{jpy_wacc_ref})*(1+{us10})/(1+{jgb})-1"
        ws.cell(row=row, column=3, value=formula)
        ws.cell(row=row, column=3).number_format = '0.00%'
        self.styler.style_data_cell(ws, row, 1, 'derived')
        self.styler.style_data_cell(ws, row, 3, 'derived')
        self.cell_refs['nippon_usd_wacc'] = f"$C${row}"
        row += 1

        # WACC Advantage
        ws.cell(row=row, column=1, value="WACC Advantage (bps)")
        formula = f"=({uss_wacc}-{self.cell_refs['nippon_usd_wacc']})*10000"
        ws.cell(row=row, column=3, value=formula)
        ws.cell(row=row, column=3).number_format = '#,##0'
        self.styler.style_data_cell(ws, row, 1, 'derived')
        self.styler.style_data_cell(ws, row, 3, 'derived')

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30

    def _create_projections_sheet(self, wb: Workbook):
        """Create projections sheet with FCF data (values from model)"""
        ws = wb.create_sheet("Projections")

        consolidated = self.analysis['consolidated']

        # Title
        ws['A1'] = "FINANCIAL PROJECTIONS"
        ws['A1'].font = self.styler.title_font
        ws['A2'] = "Note: These projections are calculated by the Python model. DCF sheets use these values with Excel formulas."
        ws['A2'].font = Font(italic=True, size=9)

        row = 4

        # Headers
        headers = ['Metric'] + [str(y) for y in self.years]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # Projection data
        metrics = [
            ('Revenue', 'Revenue', '$#,##0'),
            ('EBITDA', 'Total_EBITDA', '$#,##0'),
            ('EBITDA Margin', 'EBITDA_Margin', '0.0%'),
            ('D&A', 'DA', '$#,##0'),
            ('NOPAT', 'NOPAT', '$#,##0'),
            ('Gross Cash Flow', 'Gross_CF', '$#,##0'),
            ('CapEx', 'Total_CapEx', '$#,##0'),
            ('Change in NWC', 'Delta_WC', '$#,##0'),
            ('Free Cash Flow', 'FCF', '$#,##0'),
        ]

        # Track row positions for each metric
        metric_rows = {}

        for metric_name, col_name, fmt in metrics:
            ws.cell(row=row, column=1, value=metric_name)
            self.styler.style_data_cell(ws, row, 1)

            # Track this metric's row
            metric_rows[col_name] = row

            for col_idx, year in enumerate(self.years, 2):
                year_data = consolidated[consolidated['Year'] == year]
                value = year_data[col_name].values[0]
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.number_format = fmt

                # Color FCF
                if col_name == 'FCF':
                    value_type = 'positive' if value > 0 else 'negative' if value < 0 else None
                    self.styler.style_data_cell(ws, row, col_idx, value_type)
                else:
                    self.styler.style_data_cell(ws, row, col_idx)
            row += 1

        # Store row references for use by other sheets
        self.fcf_row = metric_rows['FCF']
        self.ebitda_row = metric_rows['Total_EBITDA']

        # Column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(self.years) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_dcf_formula_sheet(self, wb: Workbook, perspective: str):
        """Create DCF sheet with Excel formulas

        Args:
            perspective: 'USS' or 'Nippon'
        """
        ws = wb.create_sheet(f"DCF - {perspective}")

        is_uss = perspective == 'USS'
        wacc_ref = self.cell_refs.get('uss_wacc', '$C$5') if is_uss else self.cell_refs.get('nippon_usd_wacc', '$C$28')
        tg_ref = self.cell_refs.get('terminal_growth', '$C$6')
        em_ref = self.cell_refs.get('exit_multiple', '$C$7')

        # Title
        ws['A1'] = f"{perspective.upper()} DCF VALUATION"
        ws['A1'].font = self.styler.title_font
        ws['A2'] = f"WACC: =Inputs!{wacc_ref}" if is_uss else f"WACC: =Inputs!{wacc_ref} (IRP-Adjusted)"
        ws['A2'].font = self.styler.subtitle_font

        row = 4

        # === FREE CASH FLOW SECTION ===
        self.styler.style_section_header(ws, row, "FREE CASH FLOW", 1, self.num_years + 1)
        row += 1

        # Headers
        headers = ['Item'] + [str(y) for y in self.years]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # FCF row - reference Projections sheet
        ws.cell(row=row, column=1, value="Free Cash Flow")
        self.styler.style_data_cell(ws, row, 1)
        for col_idx in range(2, self.num_years + 2):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=row, column=col_idx, value=f"=Projections!{col_letter}${self.fcf_row}")
            ws.cell(row=row, column=col_idx).number_format = '$#,##0'
            self.styler.style_data_cell(ws, row, col_idx)
        fcf_row = row
        row += 1

        # Discount Factor row - FORMULA
        ws.cell(row=row, column=1, value="Discount Factor")
        self.styler.style_data_cell(ws, row, 1)
        for col_idx, year_num in enumerate(range(1, self.num_years + 1), 2):
            # =1/(1+WACC)^year
            ws.cell(row=row, column=col_idx, value=f"=1/(1+Inputs!{wacc_ref})^{year_num}")
            ws.cell(row=row, column=col_idx).number_format = '0.0000'
            self.styler.style_data_cell(ws, row, col_idx)
        df_row = row
        row += 1

        # PV of FCF row - FORMULA
        ws.cell(row=row, column=1, value="PV of FCF")
        self.styler.style_data_cell(ws, row, 1)
        for col_idx in range(2, self.num_years + 2):
            col_letter = get_column_letter(col_idx)
            # =FCF * Discount Factor
            ws.cell(row=row, column=col_idx, value=f"={col_letter}${fcf_row}*{col_letter}${df_row}")
            ws.cell(row=row, column=col_idx).number_format = '$#,##0'
            self.styler.style_data_cell(ws, row, col_idx)
        pv_fcf_row = row
        row += 2

        # === TERMINAL VALUE SECTION ===
        self.styler.style_section_header(ws, row, "TERMINAL VALUE CALCULATION", 1, 3)
        row += 1

        last_year_col = get_column_letter(self.num_years + 1)

        tv_items = [
            ("Sum of PV FCF", f"=SUM(B{pv_fcf_row}:{last_year_col}{pv_fcf_row})", '$#,##0'),
            ("", "", ""),
            ("GORDON GROWTH METHOD", "", ""),
            ("Terminal FCF (Y+1)", f"={last_year_col}{fcf_row}*(1+Inputs!{tg_ref})", '$#,##0'),
            ("Terminal Value", f"=B{row+3}/(Inputs!{wacc_ref}-Inputs!{tg_ref})", '$#,##0'),
            ("PV of Terminal Value", f"=B{row+4}*{last_year_col}{df_row}", '$#,##0'),
            ("", "", ""),
            ("EXIT MULTIPLE METHOD", "", ""),
            ("Terminal EBITDA", f"=Projections!{last_year_col}{self.ebitda_row}", '$#,##0'),
            ("Exit Multiple", f"=Inputs!{em_ref}", '0.0"x"'),
            ("Terminal Value", f"=B{row+8}*B{row+9}", '$#,##0'),
            ("PV of Terminal Value", f"=B{row+10}*{last_year_col}{df_row}", '$#,##0'),
        ]

        base_row = row
        for i, (label, formula, fmt) in enumerate(tv_items):
            ws.cell(row=row, column=1, value=label)
            if formula:
                ws.cell(row=row, column=2, value=formula)
                ws.cell(row=row, column=2).number_format = fmt
            if label:
                self.styler.style_data_cell(ws, row, 1)
                if formula:
                    self.styler.style_data_cell(ws, row, 2)
            row += 1

        sum_pv_row = base_row
        pv_tv_gordon_row = base_row + 5
        pv_tv_exit_row = base_row + 11

        row += 1

        # === ENTERPRISE VALUE SECTION ===
        self.styler.style_section_header(ws, row, "ENTERPRISE VALUE", 1, 3)
        row += 1

        ev_items = [
            ("EV - Gordon Growth", f"=B{sum_pv_row}+B{pv_tv_gordon_row}", '$#,##0'),
            ("EV - Exit Multiple", f"=B{sum_pv_row}+B{pv_tv_exit_row}", '$#,##0'),
            ("EV - Blended (Average)", f"=(B{row}+B{row+1})/2", '$#,##0'),
        ]

        for label, formula, fmt in ev_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=formula)
            ws.cell(row=row, column=2).number_format = fmt
            self.styler.style_data_cell(ws, row, 1)
            self.styler.style_data_cell(ws, row, 2, 'positive')
            row += 1

        ev_blended_row = row - 1
        row += 1

        # === EQUITY BRIDGE SECTION ===
        self.styler.style_section_header(ws, row, "EQUITY BRIDGE", 1, 3)
        row += 1

        # Get cell references for balance sheet items
        debt_ref = self.cell_refs.get('total_debt', '$C$18')
        pension_ref = self.cell_refs.get('pension', '$C$19')
        leases_ref = self.cell_refs.get('leases', '$C$20')
        cash_ref = self.cell_refs.get('cash', '$C$21')
        inv_ref = self.cell_refs.get('investments', '$C$22')
        shares_ref = self.cell_refs.get('shares', '$C$23')

        bridge_items = [
            ("Enterprise Value", f"=B{ev_blended_row}", '$#,##0'),
            ("Less: Total Debt", f"=-Inputs!{debt_ref}", '$#,##0'),
            ("Less: Pension", f"=-Inputs!{pension_ref}", '$#,##0'),
            ("Less: Leases", f"=-Inputs!{leases_ref}", '$#,##0'),
            ("Plus: Cash", f"=Inputs!{cash_ref}", '$#,##0'),
            ("Plus: Investments", f"=Inputs!{inv_ref}", '$#,##0'),
            ("= Equity Value", f"=SUM(B{row}:B{row+5})", '$#,##0'),
            ("", "", ""),
            ("Shares Outstanding", f"=Inputs!{shares_ref}", '#,##0.0'),
            ("", "", ""),
            ("SHARE PRICE", f"=MAX(0,B{row+6})/B{row+8}", '$#,##0.00'),
        ]

        for label, formula, fmt in bridge_items:
            ws.cell(row=row, column=1, value=label)
            if formula:
                ws.cell(row=row, column=2, value=formula)
                ws.cell(row=row, column=2).number_format = fmt
            if label:
                self.styler.style_data_cell(ws, row, 1)
                if formula:
                    value_type = 'positive' if label == "SHARE PRICE" else None
                    self.styler.style_data_cell(ws, row, 2, value_type)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        for col in range(3, self.num_years + 3):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_wacc_formula_sheet(self, wb: Workbook):
        """Create WACC analysis sheet with formulas"""
        ws = wb.create_sheet("WACC Analysis")

        # Title
        ws['A1'] = "WACC ANALYSIS"
        ws['A1'].font = self.styler.title_font

        # Get cell references
        uss_wacc = self.cell_refs.get('uss_wacc', '$C$5')
        jgb = self.cell_refs.get('japan_10yr', '$C$12')
        erp = self.cell_refs.get('nippon_erp', '$C$13')
        cs = self.cell_refs.get('nippon_cs', '$C$14')
        dr = self.cell_refs.get('nippon_debt_ratio', '$C$15')
        tr = self.cell_refs.get('nippon_tax_rate', '$C$16')
        us10 = self.cell_refs.get('us_10yr', '$C$11')
        nippon_wacc = self.cell_refs.get('nippon_usd_wacc', '$C$28')

        row = 3

        # USS WACC
        self.styler.style_section_header(ws, row, "USS STANDALONE WACC", 1, 3)
        row += 1

        ws.cell(row=row, column=1, value="USS WACC")
        ws.cell(row=row, column=2, value=f"=Inputs!{uss_wacc}")
        ws.cell(row=row, column=2).number_format = '0.00%'
        self.styler.style_data_cell(ws, row, 1)
        self.styler.style_data_cell(ws, row, 2)
        row += 2

        # Nippon WACC Calculation
        self.styler.style_section_header(ws, row, "NIPPON USD WACC CALCULATION (IRP Method)", 1, 3)
        row += 1

        wacc_items = [
            ("Step 1: Cost of Capital in JPY", "", ""),
            ("Japan 10-Year JGB", f"=Inputs!{jgb}", "0.00%"),
            ("+ Equity Risk Premium", f"=Inputs!{erp}", "0.00%"),
            ("= Cost of Equity (JPY)", f"=Inputs!{jgb}+Inputs!{erp}", "0.00%"),
            ("", "", ""),
            ("+ Credit Spread", f"=Inputs!{cs}", "0.00%"),
            ("= Cost of Debt (JPY)", f"=Inputs!{jgb}+Inputs!{cs}", "0.00%"),
            ("", "", ""),
            ("Equity Weight", f"=1-Inputs!{dr}", "0%"),
            ("Debt Weight", f"=Inputs!{dr}", "0%"),
            ("Tax Rate", f"=Inputs!{tr}", "0%"),
            ("", "", ""),
            ("= JPY WACC", f"=(1-Inputs!{dr})*(Inputs!{jgb}+Inputs!{erp})+Inputs!{dr}*(Inputs!{jgb}+Inputs!{cs})*(1-Inputs!{tr})", "0.00%"),
            ("", "", ""),
            ("Step 2: Convert to USD via IRP", "", ""),
            ("Formula: (1+JPY_WACC)*(1+US_10Y)/(1+JP_10Y)-1", "", ""),
            ("US 10-Year Treasury", f"=Inputs!{us10}", "0.00%"),
            ("Japan 10-Year JGB", f"=Inputs!{jgb}", "0.00%"),
            ("", "", ""),
            ("= Nippon USD WACC", f"=Inputs!{nippon_wacc}", "0.00%"),
            ("", "", ""),
            ("COMPARISON", "", ""),
            ("USS WACC", f"=Inputs!{uss_wacc}", "0.00%"),
            ("Nippon USD WACC", f"=Inputs!{nippon_wacc}", "0.00%"),
            ("WACC Advantage (bps)", f"=(Inputs!{uss_wacc}-Inputs!{nippon_wacc})*10000", "#,##0"),
        ]

        for label, formula, fmt in wacc_items:
            ws.cell(row=row, column=1, value=label)
            if formula:
                ws.cell(row=row, column=2, value=formula)
                ws.cell(row=row, column=2).number_format = fmt
            if label:
                self.styler.style_data_cell(ws, row, 1)
                if formula:
                    self.styler.style_data_cell(ws, row, 2)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 15

    def _create_equity_bridge_formula_sheet(self, wb: Workbook):
        """Create equity bridge comparison with formulas"""
        ws = wb.create_sheet("Equity Bridge")

        # Title
        ws['A1'] = "EQUITY BRIDGE COMPARISON"
        ws['A1'].font = self.styler.title_font

        row = 3

        # Headers
        headers = ["Item", "USS Standalone", "Nippon View", "Delta"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.styler.style_header_row(ws, row, 1, len(headers))
        row += 1

        # Get cell references from tracked positions
        debt_ref = self.cell_refs.get('total_debt', '$C$19')
        pension_ref = self.cell_refs.get('pension', '$C$20')
        leases_ref = self.cell_refs.get('leases', '$C$21')
        cash_ref = self.cell_refs.get('cash', '$C$22')
        inv_ref = self.cell_refs.get('investments', '$C$23')
        shares_ref = self.cell_refs.get('shares', '$C$24')

        # EV Blended is at row 27 in DCF sheets
        ev_dcf_row = 27

        # Reference DCF sheets for EV
        bridge_items = [
            ("Enterprise Value", f"='DCF - USS'!B{ev_dcf_row}", f"='DCF - Nippon'!B{ev_dcf_row}"),
            ("", "", ""),
            ("Less: Total Debt", f"=-Inputs!{debt_ref}", f"=-Inputs!{debt_ref}"),
            ("Less: Pension", f"=-Inputs!{pension_ref}", f"=-Inputs!{pension_ref}"),
            ("Less: Leases", f"=-Inputs!{leases_ref}", f"=-Inputs!{leases_ref}"),
            ("Plus: Cash", f"=Inputs!{cash_ref}", f"=Inputs!{cash_ref}"),
            ("Plus: Investments", f"=Inputs!{inv_ref}", f"=Inputs!{inv_ref}"),
            ("", "", ""),
            ("= Equity Value", "=SUM(B{start}:B{end})", "=SUM(C{start}:C{end})"),
            ("", "", ""),
            ("Shares Outstanding", f"=Inputs!{shares_ref}", f"=Inputs!{shares_ref}"),
            ("", "", ""),
            ("SHARE PRICE", "=MAX(0,B{eq})/B{sh}", "=MAX(0,C{eq})/C{sh}"),
        ]

        start_row = row
        ev_row = row

        for i, (label, uss_formula, nippon_formula) in enumerate(bridge_items):
            ws.cell(row=row, column=1, value=label)

            # Replace placeholders in formulas
            if "{start}" in uss_formula:
                uss_formula = uss_formula.format(start=start_row, end=row-1)
                nippon_formula = nippon_formula.format(start=start_row, end=row-1)
            elif "{eq}" in uss_formula:
                eq_row = row - 4  # Equity value row
                sh_row = row - 2  # Shares row
                uss_formula = uss_formula.format(eq=eq_row, sh=sh_row)
                nippon_formula = nippon_formula.format(eq=eq_row, sh=sh_row)

            if uss_formula:
                ws.cell(row=row, column=2, value=uss_formula)
                ws.cell(row=row, column=2).number_format = '$#,##0' if label != "Shares Outstanding" else '#,##0.0'
                if label == "SHARE PRICE":
                    ws.cell(row=row, column=2).number_format = '$#,##0.00'

            if nippon_formula:
                ws.cell(row=row, column=3, value=nippon_formula)
                ws.cell(row=row, column=3).number_format = '$#,##0' if label != "Shares Outstanding" else '#,##0.0'
                if label == "SHARE PRICE":
                    ws.cell(row=row, column=3).number_format = '$#,##0.00'

            # Delta column
            if label and uss_formula and label not in ["", "Shares Outstanding"]:
                ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
                ws.cell(row=row, column=4).number_format = '$#,##0' if label != "SHARE PRICE" else '$#,##0.00'

            if label:
                self.styler.style_data_cell(ws, row, 1)
                if uss_formula:
                    self.styler.style_data_cell(ws, row, 2)
                if nippon_formula:
                    self.styler.style_data_cell(ws, row, 3)
                if ws.cell(row=row, column=4).value:
                    self.styler.style_data_cell(ws, row, 4, 'positive')

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
